import math
import pandas as pd
import scipy.io
import os
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

#funciones para leer archivos .mat, limpiar los datos y guardarlos en excel


def leer_mat_a_df(ruta_archivo, log_fn=print):
    """
    lee un archivo .mat de matlab y lo convierte en un dataframe de pandas
    Args: ruta_archivo -> ruta completa al archivo .mat
            log_fn -> funcion para mostrar mensajes (por defecto print)
    Returns: dataframe con los datos del archivo, o none si hubo algun error
    """
    try:
        columnas = [
            'Ensayo', 'Lado', 'Estim Electrico', 'Latencia',
            'Tiempo Absoluto', 'Palancas Izq',
            'Palancas Der', 'Desplazamiento'
        ]
        mat_data = scipy.io.loadmat(ruta_archivo)
        matriz_datos = None
        #iteramos sobre el diccionario que devuelve loadmat buscando la primera llave que no empiece con guion bajo ya que matlab genera llaves ocultas con metadatos que no nos interesan para el analisis
        for key, value in mat_data.items():
            if not key.startswith('__'):
                matriz_datos = value
                break

        if matriz_datos is None: #verificamos si el ciclo anterior logro encontrar la matriz de datos utiles para evitar que el programa truene al intentar procesar un archivo vacio o corrupto
            log_fn(f"Error: No se encontraron datos válidos en {os.path.basename(ruta_archivo)}")
            return None

        #comprobamos estrictamente que la matriz extraida tenga exactamente las ocho columnas que definimos arriba para garantizar que el archivo mat corresponde al formato del experimento actual y no a otra prueba distinta
        if matriz_datos.shape[1] == len(columnas):
            df = pd.DataFrame(matriz_datos, columns=columnas)
            df['archivo_origen'] = os.path.basename(ruta_archivo)
            return df
        else:
            log_fn(
                f"Error en dimensiones: {os.path.basename(ruta_archivo)} "
                f"tiene {matriz_datos.shape[1]} columnas, "
                f"pero esperamos {len(columnas)}."
            )
            return None

    except Exception as e: #capturamos cualquier excepcion no prevista durante la lectura o transformacion de los datos binarios para que el script pueda registrar el fallo y continuar con el siguiente archivo sin detener toda la ejecucion del lote
        log_fn(f"Error al procesar {os.path.basename(ruta_archivo)}: {str(e)}")
        return None


def modificar_dataframe(df):
    """
    limpia el dataframe quitando columnas que no se necesitan
    y filtrando solo los ensayos donde el animal se desplazo
    Args: df -> dataframe con los datos crudos del archivo .mat
    Returns: dataframe limpio con solo los ensayos validos y columnas utiles
    """
    df = df.drop(['Tiempo Absoluto', 'Palancas Izq', 'Palancas Der', 'archivo_origen'], axis=1) #eliminamos variables irrelevantes del dataset para liberar memoria y mantener unicamente las metricas criticas de latencia y condicion de choque
    #filtramos las filas conservando exclusivamente aquellos ensayos donde el sensor registro un desplazamiento fisico real mayor a uno descartando asi falsos positivos o inactividad del sujeto
    df = df[df['Desplazamiento'] > 1].reset_index(drop=True)
    df['Ensayo'] = range(1, len(df) + 1) #reescribimos la columna de ensayos con una secuencia numerica limpia desde el uno hasta el total de filas filtradas para corregir los saltos generados al eliminar los intentos nulos
    return df


def calcular_promedios_latencia(dfs_dict, log_fn=print):
    """
    descripcion: calcula el promedio de latencia por condicion (con y sin choque) para cada archivo
    Args: dfs_dict -> diccionario con nombre de archivo como clave y dataframe como valor
            log_fn -> funcion para mostrar mensajes (por defecto print)
    Returns: dataframe con una fila por archivo y columnas de promedio seguro y riesgo
    """
    filas = []
    for nombre_archivo, df in dfs_dict.items(): #recorremos el diccionario que contiene todos los dataframes de las sesiones ya limpios para extraer las estadisticas globales de cada rata de manera iterativa
        if 'Latencia' not in df.columns or 'Estim Electrico' not in df.columns: #hacemos una validacion de seguridad para confirmar que las columnas objetivo siguen existiendo en el dataframe y asi evitar errores de calculo al promediar
            log_fn(f"Advertencia: {nombre_archivo} no tiene las columnas requeridas.")
            continue
        promedio_con = df.loc[df['Estim Electrico'] == 1, 'Latencia'].mean() #filtramos los registros donde la variable de choque electrico esta activa y calculamos la media aritmetica de sus tiempos de reaccion
        promedio_sin = df.loc[df['Estim Electrico'] == 0, 'Latencia'].mean() #aislamos los intentos catalogados como seguros donde no hubo amenaza y promediamos sus latencias para posteriormente poder compararlos con el grupo de riesgo
        filas.append({ #armamos un nuevo diccionario temporal con los resultados estadisticos redondeados a un decimal o asignamos un cero por defecto si el calculo arrojo un valor nulo antes de agregarlo a la lista de resumen global
            'Nombre de archivo': nombre_archivo,
            'Promedio Seguro': round(promedio_sin, 1) if pd.notna(promedio_sin) else 0,
            'Promedio Riesgo': round(promedio_con, 1) if pd.notna(promedio_con) else 0,
        })
    return pd.DataFrame(filas)


def guardar_excel(ruta_guardado, dfs_mat, log_fn=print):
    """
    guarda todos los dataframes en un archivo excel con formato
            la primera hoja tiene los promedios generales y las demas tienen los datos por rata
    Args: ruta_guardado -> ruta donde se va a guardar el archivo .xlsx
            dfs_mat -> diccionario con nombre de archivo como clave y dataframe como valor
            log_fn -> funcion para mostrar mensajes (por defecto print)
    Returns: true si se guardo bien, false si hubo algun error
    """
    try:
        with pd.ExcelWriter(ruta_guardado, engine='openpyxl') as writer: #iniciamos un manejador de contexto con openpyxl que nos permitira escribir multiples pestañas en un solo documento de excel de forma segura y asegurando que se libere el archivo de la memoria al terminar
            df_promedios = calcular_promedios_latencia(dfs_mat, log_fn)

            #se calcula el promedio general de todas las ratas
            col_seguro = df_promedios['Promedio Seguro']
            col_riesgo = df_promedios['Promedio Riesgo']

            promedio_seguro = round(col_seguro.mean(), 1)
            promedio_riesgo = round(col_riesgo.mean(), 1)

            #se calcula el sem como desviacion estandar dividida entre raiz de n
            n_seguro = col_seguro.count()
            n_riesgo = col_riesgo.count()
            sem_seguro = round(col_seguro.std() / math.sqrt(n_seguro), 3) if n_seguro > 1 else 0 #calculamos el error estandar de la media para los ensayos seguros validando primero que tengamos mas de un dato para que la operacion de desviacion estandar no genere una division por cero
            sem_riesgo = round(col_riesgo.std() / math.sqrt(n_riesgo), 3) if n_riesgo > 1 else 0 #aplicamos el mismo calculo de varianza estadistica para la condicion de riesgo asegurandonos de tener una muestra minima indispensable para dar validez al resultado

            #se agregan filas de promedio y sem al final de la tabla
            idx_sep      = len(df_promedios) + 2
            idx_promedio = len(df_promedios) + 3
            idx_sem      = len(df_promedios) + 4

            df_promedios.loc[idx_sep,      'Nombre de archivo'] = '' #insertamos una fila vacia usando su indice posicional especifico para crear una separacion visual limpia entre el listado individual de ratas y los resultados globales calculados
            df_promedios.loc[idx_promedio, 'Nombre de archivo'] = 'Promedio'
            df_promedios.loc[idx_promedio, 'Promedio Seguro']   = promedio_seguro
            df_promedios.loc[idx_promedio, 'Promedio Riesgo']   = promedio_riesgo
            df_promedios.loc[idx_sem,      'Nombre de archivo'] = 'SEM'
            df_promedios.loc[idx_sem,      'Promedio Seguro']   = sem_seguro
            df_promedios.loc[idx_sem,      'Promedio Riesgo']   = sem_riesgo

            #se escribe primero la hoja de promedios para que quede de primera
            nombre_hoja_prom = 'Promedios Latencia'
            df_promedios.to_excel(writer, sheet_name=nombre_hoja_prom, index=False)
            ws_prom = writer.sheets[nombre_hoja_prom]
            for row in ws_prom.iter_rows(): #recorremos la matriz completa de la recien creada hoja de promedios para aplicar un estilo de fuente en negrita a absolutamente todas sus celdas logrando asi que resalte inmediatamente como el sumario principal al abrir el reporte
                for cell in row:
                    cell.font = Font(bold=True)
            for col_idx, column in enumerate(df_promedios.columns, start=1): #iteramos a traves de cada columna generada para examinar su contenido y calcular de forma dinamica la longitud del texto mas largo contenido en cualquiera de sus celdas
                col_letter = get_column_letter(col_idx)
                max_length = len(str(column))
                for cell in ws_prom[col_letter]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws_prom.column_dimensions[col_letter].width = max_length + 2 #ajustamos el ancho de la columna actual basandonos en la longitud maxima detectada y le sumamos un pequeño margen para asegurar que ningun dato numerico ni encabezado quede truncado a la vista

            #se agrega una hoja por cada rata con sus datos completos
            for nombre_archivo, df_final in dfs_mat.items(): #procesamos el diccionario principal separando cada dataframe individual para inyectarlo en su propia pestaña dedicada utilizando como nombre base el identificador del archivo original de procedencia
                nombre_hoja = nombre_archivo.replace('.mat', '')[:31] #limpiamos la extension del nombre del archivo y truncamos la cadena a treinta y un caracteres exactos para cumplir estrictamente con el limite maximo permitido por la especificacion de excel para nombrar pestañas
                df_final.to_excel(writer, sheet_name=nombre_hoja, index=False)
                worksheet = writer.sheets[nombre_hoja]

                #solo se pone negrita en el encabezado de cada hoja
                for cell in worksheet[1]: #seleccionamos de manera exclusiva la primera fila del documento que corresponde a los titulos de las variables experimentales y le aplicamos un peso de fuente en negrita para diferenciarlo visualmente de los datos crudos
                    cell.font = Font(bold=True)

                for col_idx, column in enumerate(df_final.columns, start=1): #repetimos nuestro algoritmo de autoajuste de columnas iterando por cada variable especifica de la rata actual para medir las cadenas de texto y expandir el ancho horizontal de las celdas evitando el traslape de la informacion visual
                    col_letter = get_column_letter(col_idx)
                    max_length = len(str(column))
                    for cell in worksheet[col_letter]:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    worksheet.column_dimensions[col_letter].width = max_length + 2

        log_fn(f"Archivo guardado en {ruta_guardado}")
        return True

    except Exception as e: #implementamos un bloque global de captura de excepciones para asegurar que cualquier fallo inesperado durante la escritura o formateo del archivo binario no colapse el script principal retornando en su lugar una bandera booleana de falso
        log_fn(f"Error crítico al guardar: {str(e)}")
        return False