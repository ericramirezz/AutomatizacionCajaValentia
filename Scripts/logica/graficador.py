import os
import math
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


#funciones de apoyo que se usan para calcular cosas antes de graficar

def _sem(valores):
    """
    calcula el error estandar de una lista de numeros validando el tamaño de la muestra
    Args: valores -> lista con las latencias individuales (list)
    Return: error estandar de las latencias calculadas (float)
    """
    n = len(valores)
    if n < 2:
        #si hay menos de dos valores no tiene sentido calcularlo porque no hay varianza
        return 0.0
    s = pd.Series(valores)
    return float(s.std() / math.sqrt(n))


def _tercio_split(serie):
    """
    divide los ensayos en tres bloques lo mas iguales posible dando prioridad de tamaño al primer grupo
    Args: serie -> valores de latencia crudos (series)
    Return: lista con tres series divididas una por cada bloque temporal (list)
    """
    n = len(serie)
    if n == 0: #si no hay datos devolvemos tres series completamente vacias
        return [pd.Series(dtype=float)] * 3 
    g1 = math.ceil(n / 3) #sacamos el tamaño del primer grupo redondeando hacia arriba
    rem = n - g1 #calculamos lo que queda para los otros dos grupos
    g2 = math.ceil(rem / 2) if rem > 0 else 0 #definimos el segundo grupo dividiendo el remanente
    return [ 
        serie.iloc[:g1].reset_index(drop=True), #cortamos la primera parte
        serie.iloc[g1: g1 + g2].reset_index(drop=True), #cortamos la parte del medio
        serie.iloc[g1 + g2:].reset_index(drop=True), #cortamos el resto que queda al final
    ]


def _x_tercio(dia, tercio):
    """
    determina la posicion horizontal exacta donde se dibujara cada punto en la grafica
    Args: dia -> numero del dia de la sesion (int)
          tercio -> numero del bloque temporal especifico (int)
    Return: posicion en el eje x del punto a graficar (float)
    """
    offsets = {1: -0.25, 2: 0.0, 3: 0.25} #definimos los ajustes de posicion lateral para separar visualmente cada tercio
    return dia + offsets[tercio] #regresamos el valor base del dia sumado al ajuste de su tercio correspondiente


#funcion principal que lee los archivos procesa y arma el renderizado de la grafica

def generar_grafica(archivos, log_fn=print):
    """
    lee los archivos excel de cada dia calcula promedios por bloque temporal y genera la grafica de dispersion con barras de error
    Args: archivos -> lista de rutas absolutas a los archivos .xlsx (list)
          log_fn -> funcion inyectada para mostrar mensajes en la consola (callable)
    Return: tupla estructurada con la figura grafica el dataframe de resumen y los diccionarios de datos crudos
    """
    filas_resumen = []
    xs_seguro, ys_seguro, sems_seguro = [], [], []
    xs_riesgo,  ys_riesgo,  sems_riesgo  = [], [], []

    dias_ratas_raw = {}

    for dia_idx, ruta in enumerate(archivos, start=1): #iniciamos el bucle principal iterando la lista de rutas para procesar cada archivo excel de la sesion diaria
        try:
            archivo_excel = pd.ExcelFile(ruta)

            ratas_del_dia = {}
            rata_pos = 0

            for hoja in archivo_excel.sheet_names: #iteramos sobre todas las pestañas del libro de excel asumiendo que cada una contiene los ensayos completos de una rata
                if hoja == 'Promedios Latencia': #brincamos la hoja de resumen automatico porque nuestro algoritmo calcula sus metricas desde los datos crudos
                    continue
                df_tmp = archivo_excel.parse(hoja)
                if 'Latencia' not in df_tmp.columns or 'Estim Electrico' not in df_tmp.columns: #descartamos cualquier hoja que no tenga las variables dependientes e independientes necesarias
                    continue

                rata_pos += 1
                seg = df_tmp.loc[df_tmp['Estim Electrico'] == 0,
                                'Latencia'].reset_index(drop=True)
                rie = df_tmp.loc[df_tmp['Estim Electrico'] == 1,
                                'Latencia'].reset_index(drop=True)

                ratas_del_dia[rata_pos] = {
                    'seguro': _tercio_split(seg),
                    'riesgo': _tercio_split(rie),
                }

            dias_ratas_raw[dia_idx] = ratas_del_dia

            if not ratas_del_dia: #notificamos un error sin detener la ejecucion si la estructura de datos quedo vacia indicando que ninguna pestaña paso los filtros
                log_fn(f"Día {dia_idx}: no se encontraron datos válidos.")
                continue

            for t_idx in range(3): #establecemos un ciclo de tres iteraciones para segmentar y procesar la evolucion del aprendizaje en bloques inicial medio y final
                tercio_num = t_idx + 1

                medias_seg = []
                medias_rie = []

                for rata_data in ratas_del_dia.values(): #recorremos los diccionarios de latencias extraidas para calcular las medias aritmeticas individuales por condicion
                    grupo_seg = rata_data['seguro'][t_idx]
                    grupo_rie = rata_data['riesgo'][t_idx]
                    if len(grupo_seg) > 0: #verificamos que el sujeto haya ejecutado intentos seguros en este tercio para evitar un fallo de division por cero al promediar
                        medias_seg.append(float(grupo_seg.mean()))
                    if len(grupo_rie) > 0: #nos cercioramos de que existan ensayos bajo amenaza antes de incluir la latencia en la lista de estadistica poblacional
                        medias_rie.append(float(grupo_rie.mean()))

                if medias_seg: #calculamos el promedio global y su varianza estadistica si logramos recolectar medias seguras validas entre todas las ratas
                    prom_seg = round(float(pd.Series(medias_seg).mean()), 2)
                    sem_seg  = round(_sem(medias_seg), 3)
                else:
                    prom_seg, sem_seg = None, None

                if medias_rie: #evaluamos la existencia de datos de riesgo para generar el estimador poblacional de la latencia en condiciones aversivas
                    prom_rie = round(float(pd.Series(medias_rie).mean()), 2)
                    sem_rie  = round(_sem(medias_rie), 3)
                else:
                    prom_rie, sem_rie = None, None

                x = _x_tercio(dia_idx, tercio_num)

                if prom_seg is not None and prom_seg != 0: #filtramos los resultados nulos asegurando que enviamos unicamente coordenadas fidedignas al motor de graficado
                    xs_seguro.append(x)
                    ys_seguro.append(prom_seg)
                    sems_seguro.append(sem_seg)

                if prom_rie is not None and prom_rie != 0: #validamos que el promedio de riesgo sea util descartando ceros que distorsionarian la percepcion de la tarea
                    xs_riesgo.append(x)
                    ys_riesgo.append(prom_rie)
                    sems_riesgo.append(sem_rie)

                filas_resumen.append({
                    'Día':         dia_idx,
                    'Bloque':      tercio_num,
                    'Prom Seguro': prom_seg,
                    'SEM Seguro':  sem_seg,
                    'Prom Riesgo': prom_rie,
                    'SEM Riesgo':  sem_rie,
                })

            total_ensayos = sum(
                sum(len(rd[cond][t]) for cond in ('seguro', 'riesgo') for t in range(3))
                for rd in ratas_del_dia.values()
            )
            log_fn(
                f"Día {dia_idx}: {os.path.basename(ruta)} — "
                f"{total_ensayos} ensayos procesados ({rata_pos} rata(s))."
            )

        except Exception as e:
            log_fn(f"Error procesando {os.path.basename(ruta)}: {str(e)}")

    if not xs_seguro and not xs_riesgo: #verificacion final de seguridad para abortar el dibujado si todos los archivos procesados regresaron arreglos de coordenadas vacios
        log_fn("No hay datos válidos para graficar.")
        return None, None, None

    num_dias = len(archivos)
    
    #calculamos dinamicamente el ancho de la figura multiplicando 1.1 pulgadas por dia garantizando una distancia fija
    #establecemos un ancho minimo de 5 pulgadas para que el titulo y la leyenda siempre quepan comodamente aunque sea un solo dia
    ancho_fig = max(5.0, 11.0 * (num_dias / 10.0))
    fig, ax = plt.subplots(figsize=(ancho_fig, 6))
    # -------------------------
    
    err_kw = dict(capsize=0, capthick=1.5, elinewidth=1.5)

    if xs_seguro: #dibujamos los marcadores circulares y barras de error de la condicion segura restaurando el formato sin guion
        ax.errorbar(xs_seguro, ys_seguro, yerr=sems_seguro,
                    fmt='o', color='green', markersize=8, alpha=0.85,
                    label='No - Conflict', **err_kw)
        
        #aislamos los dias detectados para trazar una linea continua exclusivamente entre los tres tercios de una misma sesion
        dias_seg = sorted(list(set(round(x) for x in xs_seguro)))
        for d in dias_seg:
            x_linea = [x for x in xs_seguro if round(x) == d]
            y_linea = [y for x, y in zip(xs_seguro, ys_seguro) if round(x) == d]
            ax.plot(x_linea, y_linea, '-', color='green', alpha=0.85)

    if xs_riesgo: #dibujamos los puntos negros y barras de error para el grupo de conflicto aversivo sin unirlos todos juntos
        ax.errorbar(xs_riesgo, ys_riesgo, yerr=sems_riesgo,
                    fmt='o', color='black', markersize=8, alpha=0.85,
                    label='Conflict', **err_kw)
        
        #recorremos cada dia de riesgo independientemente dibujando lineas que conectan unicamente los bloques internos del dia
        dias_rie = sorted(list(set(round(x) for x in xs_riesgo)))
        for d in dias_rie:
            x_linea = [x for x in xs_riesgo if round(x) == d]
            y_linea = [y for x, y in zip(xs_riesgo, ys_riesgo) if round(x) == d]
            ax.plot(x_linea, y_linea, '-', color='black', alpha=0.85)
    
    any_seg = len(xs_seguro) > 0
    any_rie = len(xs_riesgo) > 0
    
    if any_seg and any_rie: #evaluamos la coexistencia de condiciones para asignar el titulo de discriminacion y ajustar el limite del eje y a 190 segundos
        titulo = 'Discrimination'
        ax.set_ylim(0, 190)
    elif not any_seg: #configuramos el titulo para cruces bajo amenaza pura y limitamos el eje y a 130 segundos para no cortar las barras
        titulo = 'Threat crossings'
        ax.set_ylim(0, 130)
    else: #ajustamos el contexto a recompensa natural sin amenaza con un tope de 110 segundos en el eje y
        titulo = 'Reward crossings'
        ax.set_ylim(0, 110)

    ax.set_title(titulo, fontsize=14, fontweight='bold')
    ax.set_xlabel('Days (block of trial)', fontsize=12)
    ax.set_ylabel('Latencies (s)', fontsize=12)
    
    #restringimos el marco visual al numero real de dias procesados para recortar la imagen y evitar el espacio en blanco sobrante
    ax.set_xlim(0.5, num_dias + 0.5) 
    
    #ponemos las marcas numericas precisas sin generar numeros extra
    ax.set_xticks(range(1, num_dias + 1)) 
    ax.set_xticklabels([str(d) for d in range(1, num_dias + 1)], fontsize=10) 
    # -------------------------
    
    ax.yaxis.set_major_locator(ticker.MultipleLocator(20)) #obligamos al eje vertical a trazar sus numeros y lineas guia estrictamente en saltos de 20 en 20
    
    ax.legend(fontsize=11)
    ax.grid(False)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    fig.tight_layout()

    df_resumen = pd.DataFrame(filas_resumen)
    log_fn("Gráfica generada correctamente.")
    return fig, df_resumen, dias_ratas_raw

#funcion que escribe la hoja con los datos crudos de cada rata separados por bloque

def _escribir_hoja_raw(ws, dias_ratas_raw, num_dias):
    """
    estructura y llena la pestaña de datos crudos insertando las latencias individuales de cada rata organizadas por bloques y dias
    Args: ws -> referencia al objeto de la hoja de calculo donde se escribiran los datos (worksheet)
          dias_ratas_raw -> estructura de diccionarios anidados conteniendo todas las latencias procesadas (dict)
          num_dias -> cantidad entera que indica el limite de iteraciones para las columnas (int)
    Return: no regresa valores directamente sino que muta el documento de excel cargado en memoria
    """
    nombres_tercio = ['Primer bloque', 'Segundo bloque', 'Tercer bloque']

    header = ['']
    for d in range(1, num_dias + 1): #iteramos a traves de los dias totales para generar dinamicamente las cabeceras emparejando condicion segura y de riesgo
        header.append(f'Seguro_d{d}')
        header.append(f'Riesgo_d{d}')
    for col_idx, h in enumerate(header, start=1): #recorremos la lista de cabeceras aplicando inmediatamente un formato de texto en negritas a la celda destino
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)

    max_ratas = 0
    for ratas in dias_ratas_raw.values(): #exploramos los registros diarios para aislar el numero identificador maximo de rata registrado en la sesion completa
        if ratas: #evaluamos que el diccionario diario contenga elementos para evitar excepciones fatales al calcular el limite maximo
            max_ratas = max(max_ratas, max(ratas.keys()))

    current_row = 2

    for rata_idx in range(1, max_ratas + 1): #iniciamos el ciclo de escritura iterando sobre el censo total de ratas para generar los bloques individuales de datos
        #escribimos la etiqueta de identificacion de la rata en la primera columna con fuente resaltada
        cell = ws.cell(row=current_row, column=1, value=f'r{rata_idx}')
        cell.font = Font(bold=True)
        current_row += 1

        for t_idx in range(3): #establecemos un bucle interno de tres ciclos para organizar la informacion del sujeto en bloques temporales cronologicos
            segs_por_dia = []
            ries_por_dia = []

            for dia_idx in range(1, num_dias + 1): #recorremos los dias experimentales buscando en la base de datos la informacion especifica de esta rata
                ratas     = dias_ratas_raw.get(dia_idx, {})
                rata_data = ratas.get(rata_idx)
                if rata_data: #comprobamos la existencia de un registro valido para extraer las series y convertirlas nativamente a listas de python
                    segs_por_dia.append(rata_data['seguro'][t_idx].tolist())
                    ries_por_dia.append(rata_data['riesgo'][t_idx].tolist())
                else: #insertamos listas vacias si el sujeto no corrio la prueba ese dia manteniendo asi la alineacion estructural de las celdas
                    segs_por_dia.append([])
                    ries_por_dia.append([])

            max_vals = max(
                max((len(s) for s in segs_por_dia), default=0),
                max((len(r) for r in ries_por_dia), default=0),
                1,
            )

            for val_row in range(max_vals): #iteramos el numero maximo de ensayos detectados generando filas suficientes para acomodar toda la informacion
                row = current_row + val_row
                if val_row == 0: #identificamos la fila de arranque de cada bloque para colocar el subtitulo descriptivo en la primera columna libre
                    ws.cell(row=row, column=1, value=f'  {nombres_tercio[t_idx]}')

                for dia_idx in range(1, num_dias + 1): #reiteramos sobre los dias para calcular indices de columna y volcar los valores numericos individuales de latencia
                    seg_col  = 2 + (dia_idx - 1) * 2
                    rie_col  = seg_col + 1
                    seg_vals = segs_por_dia[dia_idx - 1]
                    rie_vals = ries_por_dia[dia_idx - 1]

                    if val_row < len(seg_vals): #validamos que el iterador no supere la longitud de la lista de ensayos para escribir el valor redondeado a cuatro decimales
                        ws.cell(row=row, column=seg_col,
                                value=round(float(seg_vals[val_row]), 4))
                    if val_row < len(rie_vals): #verificamos cuidadosamente los limites de la matriz antes de extraer el numero de riesgo e insertarlo en su celda respectiva
                        ws.cell(row=row, column=rie_col,
                                value=round(float(rie_vals[val_row]), 4))

            current_row += max_vals
            current_row += 1  #agregamos un salto de linea esteril para separar esteticamente los bloques temporales

        current_row += 1  #sumamos una fila blanca extra al terminar de procesar un sujeto para aislarlo del siguiente

    ws.column_dimensions['A'].width = 20
    for d in range(1, num_dias + 1): #calculamos dinamicamente las letras de columna correspondientes a los dias para aplicar anchos uniformes que prevengan traslapes de texto
        ws.column_dimensions[get_column_letter(2 + (d - 1) * 2)].width = 14
        ws.column_dimensions[get_column_letter(3 + (d - 1) * 2)].width = 14

def _escribir_hoja_promedios(ws, dias_ratas_raw, num_dias):
    """
    consolida y vacia en una hoja dedicada los promedios individuales por dia de cada rata finalizando con la media general
    Args: ws -> referencia a la hoja de excel destino generada en memoria (worksheet)
          dias_ratas_raw -> estructura de datos conteniendo todos los tiempos crudos indexados (dict)
          num_dias -> conteo total de dias de grabacion procesados (int)
    Return: opera mediante efectos secundarios alterando el objeto de excel directamente
    """

    header = ['Rata']
    for d in range(1, num_dias + 1): #iteramos a lo largo del experimento para construir la fila de encabezados añadiendo pares de columnas de condiciones
        header.append(f'Seguro Día {d}')
        header.append(f'Riesgo Día {d}')
    col_prom_seg = len(header) + 1
    col_prom_rie = col_prom_seg + 1
    header.append('Prom Seguro')
    header.append('Prom Riesgo')

    for col_idx, h in enumerate(header, start=1): #escribimos la lista de cabeceras en la primera fila de la hoja de excel e inmediatamente le inyectamos tipografia negrita
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)

    max_ratas = 0
    for ratas in dias_ratas_raw.values(): #rastreamos iterativamente el historial de diccionarios para definir cuantas filas base necesitaremos asumiendo la rata maxima
        if ratas: #protegemos el calculo matematico validando que las llaves existan previniendo interrupciones por colecciones vacias
            max_ratas = max(max_ratas, max(ratas.keys()))

    for rata_idx in range(1, max_ratas + 1): #iniciamos el volcado principal procesando individualmente la linea de tiempo historica de cada sujeto experimental detectado
        row = rata_idx + 1
        cell = ws.cell(row=row, column=1, value=f'r{rata_idx}')
        cell.font = Font(bold=True)

        promedios_seg_dias = []
        promedios_rie_dias = []

        for dia_idx in range(1, num_dias + 1): #recorremos dia a dia la base de datos de esta rata especifica aislando sus tiempos para unificarlos y colapsarlos a promedios diarios
            ratas     = dias_ratas_raw.get(dia_idx, {})
            rata_data = ratas.get(rata_idx)

            seg_col = 2 + (dia_idx - 1) * 2
            rie_col = seg_col + 1

            if rata_data: #verificamos la presencia de un rastro de ensayos de la rata en la sesion actual para proceder a concatenar y limpiar valores nulos
                #juntamos los bloques del dia usando concat de pandas para poder obtener la media global de la sesion
                all_seg = pd.concat(rata_data['seguro']).dropna()
                all_rie = pd.concat(rata_data['riesgo']).dropna()

                if len(all_seg) > 0: #comprobamos la existencia de un ensayo seguro util para procesar promedios redondeados y guardarlos visualmente en su celda respectiva
                    prom_dia_seg = round(float(all_seg.mean()), 4)
                    ws.cell(row=row, column=seg_col, value=prom_dia_seg)
                    promedios_seg_dias.append(prom_dia_seg)

                if len(all_rie) > 0: #nos aseguramos de que existan eventos bajo amenaza despues del dropna para calcular la media aritmetica y alimentarla a las listas de acumulacion historica
                    prom_dia_rie = round(float(all_rie.mean()), 4)
                    ws.cell(row=row, column=rie_col, value=prom_dia_rie)
                    promedios_rie_dias.append(prom_dia_rie)

        if promedios_seg_dias: #evaluamos si la acumulacion final de medias seguras de la rata tiene contenido para resolver la media total del experimento y aplicarle formato pesado
            cell = ws.cell(row=row, column=col_prom_seg,
                        value=round(float(pd.Series(promedios_seg_dias).mean()), 4))
            cell.font = Font(bold=True)
        if promedios_rie_dias: #confirmamos si el sujeto corrio ensayos con choque electrico en algun momento de la prueba para obtener su promedio total absoluto en el reporte final
            cell = ws.cell(row=row, column=col_prom_rie,
                        value=round(float(pd.Series(promedios_rie_dias).mean()), 4))
            cell.font = Font(bold=True)

    ws.column_dimensions['A'].width = 10
    for d in range(1, num_dias + 1): #calculamos dinamicamente el identificador de cada letra de columna y expandimos su ancho horizontal garantizando la legibilidad integral del reporte
        ws.column_dimensions[get_column_letter(2 + (d - 1) * 2)].width = 15
        ws.column_dimensions[get_column_letter(3 + (d - 1) * 2)].width = 15
    ws.column_dimensions[get_column_letter(col_prom_seg)].width = 14
    ws.column_dimensions[get_column_letter(col_prom_rie)].width = 14

def guardar_excel_resumen(df_resumen, ruta_xlsx, dias_ratas_raw=None, num_dias=0, log_fn=print):
    """
    controlador final que empaqueta las tres pestañas procesadas y las escribe definitivamente en un archivo de excel fisico
    Args: df_resumen -> tabla de pandas concentrando los promedios y estadistica basica por dia (dataframe)
          ruta_xlsx -> ruta y nombre absoluto donde se generara el nuevo reporte procesado (str)
          dias_ratas_raw -> base de datos en crudo opcional para exportar desglose (dict)
          num_dias -> delimitador entero del tamaño muestral total (int)
          log_fn -> emisor de eventos para la interfaz (callable)
    Return: bandera booleana indicando el exito de la operacion de guardado en el disco (bool)
    """
    try:
        with pd.ExcelWriter(ruta_xlsx, engine='openpyxl') as writer:
            #volcamos la tabla principal de resumenes estadisticos bloque por bloque
            df_resumen.to_excel(writer, sheet_name='Resumen Bloques', index=False)
            ws_res = writer.sheets['Resumen Bloques']
            for cell in ws_res[1]:
                cell.font = Font(bold=True)
            for col_idx, col in enumerate(df_resumen.columns, start=1):
                col_letter = get_column_letter(col_idx)
                max_len = max(
                    len(str(col)),
                    *[len(str(c.value)) for c in ws_res[col_letter]
                    if c.value is not None]
                )
                ws_res.column_dimensions[col_letter].width = max_len + 3

            if dias_ratas_raw and num_dias > 0:
                ws_raw = writer.book.create_sheet(title='Desglose Bloques')
                _escribir_hoja_raw(ws_raw, dias_ratas_raw, num_dias)

            if dias_ratas_raw and num_dias > 0:
                ws_prom = writer.book.create_sheet(title='Promedio Rata')
                _escribir_hoja_promedios(ws_prom, dias_ratas_raw, num_dias)

        log_fn(f"Excel de resumen guardado en: {ruta_xlsx}")
        return True
    except Exception as e:
        log_fn(f"Error guardando Excel: {str(e)}")
        return False